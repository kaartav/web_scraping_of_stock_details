import requests

import json
e=requests.get("https://www.alphavantage.co/query?function=FX_DAILY&from_symbol=EUR&to_symbol=USD&apikey=7DO4PV7MPHHPITOC")
#print(e.text)
k=json.loads(e.text)


startdate=int(input(f"enter the start day in augest:"))

enddate=int(input(f"enter the end day in augest:"))

###--------------------has some error mostly d=from api site-------------------------

from openpyxl import Workbook,load_workbook
wb=load_workbook("stocks.xlsx")
ws=wb.active
ws['b1']=" opening values"
ws['c1']="highest values"
ws['d1']="lowest values"
ws['e1']="closing values"

i=0
for row in range(2,((enddate-startdate)+3)):
   
    
    ws[f'a{row}'].value = f"{startdate+i} /08/2023"
    
    if(startdate+i<10):
        jim =(f"{(startdate+i):02}")

    if (f"2023-08-{jim}") in  k["Time Series FX (Daily)"]:
        ws[f'b{row}'].value = (k["Time Series FX (Daily)"][f"2023-08-{jim}"]["1. open"])
        ws[f'c{row}'].value = (k["Time Series FX (Daily)"][f"2023-08-{jim}"]["2. high"])
        ws[f'd{row}'].value = (k["Time Series FX (Daily)"][f"2023-08-{jim}"]["3. low"])
        ws[f'e{row}'].value = (k["Time Series FX (Daily)"][f"2023-08-{jim}"]["4. close"])
    else:
        ws[f'b{row}'].value = "not goven in api"
    i=i+1
wb.save("stocks.xlsx")